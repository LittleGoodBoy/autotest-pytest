from openpyxl import load_workbook
from untils.readYaml import baseConfig
import xlrd
tiqu_path=baseConfig.TIQU_PATH
api_case_path=baseConfig.APICASE_EXCEL_PATH

class Excel_rw():
    def __init__(self,filename):
        self.filename=filename
        self.wb = load_workbook(self.filename)  # 读取一个excel
        sheets_name = self.wb.sheetnames  # 获取所有的sheet表格名称，返回的
        self.sheet1 = self.wb[sheets_name[0]]  # 第一个sheet表格
        '''xlrd'''
        self.excel=xlrd.open_workbook(filename)
    def write_tiqu(self,rowv,colV):
        #获取有多少行
        rows=self.sheet1.max_row
        #写入第一行的值 ：变量名称
        self.sheet1.cell(row=rows+1,column=1).value=rowv
        #写入第二行的值 ：变量具体的值
        self.sheet1.cell(row=rows+1,column=2).value=colV
        self.wb.save(self.filename)

    def read_tiqu(self,name):
        #循环第一列的值
        rows=self.sheet1.max_row  #4
        for i in range(2,rows+1): #i=2,3,4
            clo_value=self.sheet1.cell(row=i, column=1).value
            if clo_value==name:
                data=self.sheet1.cell(row=i, column=2).value
                return data
        #如果值等于name，就取这行的第二列的值

    def read_api_cases(self):
        sheets=self.excel.sheet_names() #获取有多少个表格，返回的是列表
        print('表格数量',sheets)
        casesInfo = []  # 所有用例的数据的汇总是列表，列表的元素为字典，字典代表每个用例的信息
        for s in range(len(sheets)):
            sheet = self.excel.sheet_by_index(s)  # 第几个sheet表格
            rows, cols = sheet.nrows, sheet.ncols  # 获取行数和列数
            headers = sheet.row_values(0)  # excel的头部,固定的
            # print(rows,cols,headers,)
            for i in range(1, rows):  # 每次循环之后，会往列表里面添加一个新的字典，也就是新的用例
                caseInfo = {}  # 字典赋值的格式：{key头部信息第一列，value对应每一行第一个值}
                for j in range(0, cols):  # 等有多少列全部循环之后，添加字典
                    caseInfo[headers[j]] = sheet.row_values(i)[j]
                casesInfo.append(caseInfo)
        #print(casesInfo)
        return casesInfo
    # ['接口用例名称', '接口url', '请求方式', '请求头信息', '请求入参', '提取变量', '检查点']
    # ['登录1',‘login’,'post','','']
    # ['登录2',‘login’,'post','',]

    #{"接口用例名称": "登录", "接口入参": "{}"}

if __name__ == '__main__':
    a=Excel_rw(filename=api_case_path)
    #a.write_tiqu('a','b')
    #print(a.read_tiqu('zhuangtaima'))
    a.read_api_cases()